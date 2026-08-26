import os, io, base64, json, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, render_template_string, send_file, redirect, session, Response
from datetime import datetime, timedelta
import qrcode

app = Flask(__name__)
app.secret_key = "utesa_carnetizacion_2024"

CARPETA = "/tmp/documentos"
HISTORIAL_FILE = "/tmp/historial.json"
TURNO_FILE = "/tmp/turno.json"
ESTADO_FILE = "/tmp/estado.json"
CONFIG_FILE = "/tmp/config.json"
os.makedirs(CARPETA, exist_ok=True)

CLAVE_DEFAULT = "utesa2026"
EXTS = {"pdf","jpg","jpeg","png","doc","docx"}
MAX_MB = 25

UTESA_LOGO = "https://upload.wikimedia.org/wikipedia/commons/thumb/2/2e/UTESA_logo.png/200px-UTESA_logo.png"

def ext_ok(n):
    return "." in n and n.rsplit(".",1)[1].lower() in EXTS

def es_imagen(n):
    return "." in n and n.rsplit(".",1)[1].lower() in {"jpg","jpeg","png"}

def generar_qr(url):
    qr = qrcode.QRCode(box_size=8, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()

def cargar_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE) as f:
            return json.load(f)
    return {"clave": CLAVE_DEFAULT, "mensaje_dia": ""}

def guardar_config(c):
    with open(CONFIG_FILE,"w") as f:
        json.dump(c, f, ensure_ascii=False)

def get_clave():
    return cargar_config().get("clave", CLAVE_DEFAULT)

def cargar_historial():
    if os.path.exists(HISTORIAL_FILE):
        with open(HISTORIAL_FILE) as f:
            return json.load(f)
    return []

def guardar_historial(h):
    with open(HISTORIAL_FILE,"w") as f:
        json.dump(h, f, ensure_ascii=False)

def cargar_turno():
    if os.path.exists(TURNO_FILE):
        with open(TURNO_FILE) as f:
            return json.load(f).get("actual", 0)
    return 0

def siguiente_turno():
    actual = cargar_turno()
    nuevo = (actual % 50) + 1
    with open(TURNO_FILE,"w") as f:
        json.dump({"actual": nuevo}, f)
    return nuevo

def cargar_estado():
    if os.path.exists(ESTADO_FILE):
        with open(ESTADO_FILE) as f:
            return json.load(f)
    return {"turno_actual": 0, "atendidos_hoy": 0, "fecha_hoy": "", "stats": {}}

def guardar_estado(e):
    with open(ESTADO_FILE,"w") as f:
        json.dump(e, f, ensure_ascii=False)

def get_estado():
    e = cargar_estado()
    hoy = datetime.now().strftime("%Y-%m-%d")
    if e.get("fecha_hoy") != hoy:
        if e.get("fecha_hoy"):
            stats = e.get("stats", {})
            stats[e["fecha_hoy"]] = e.get("atendidos_hoy", 0)
            e["stats"] = stats
        e["atendidos_hoy"] = 0
        e["fecha_hoy"] = hoy
        guardar_estado(e)
    return e

def stats_semana():
    e = cargar_estado()
    stats = e.get("stats", {})
    hoy = datetime.now()
    dias = []
    for i in range(6, -1, -1):
        d = (hoy - timedelta(days=i)).strftime("%Y-%m-%d")
        label = (hoy - timedelta(days=i)).strftime("%d/%m")
        count = stats.get(d, 0)
        if d == hoy.strftime("%Y-%m-%d"):
            count = e.get("atendidos_hoy", 0)
        dias.append({"fecha": label, "count": count})
    return dias

LOGIN = """<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Carnetización - Acceso</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#f0f2f5;display:flex;align-items:center;justify-content:center;min-height:100vh;padding:20px}
.card{background:white;width:100%;max-width:380px;padding:32px;border-radius:16px;box-shadow:0 4px 16px rgba(0,0,0,.1);text-align:center}
img.logo{width:80px;margin-bottom:12px}
h2{color:#1a1a2e;margin-bottom:8px;font-size:20px}p{color:#777;font-size:14px;margin-bottom:24px}
input{width:100%;padding:12px;border:1px solid #ccc;border-radius:8px;font-size:15px;margin-bottom:14px}
button{width:100%;padding:13px;background:#1a1a2e;color:white;border:none;border-radius:8px;font-size:16px;cursor:pointer}
.err{color:#c0392b;font-size:13px;margin-bottom:12px}
</style></head><body><div class="card">
<img class="logo" src="{{ logo }}" onerror="this.style.display='none'">
<h2>Carnetización UTESA</h2><p>Panel de la encargada</p>
{% if error %}<div class="err">Contraseña incorrecta.</div>{% endif %}
<form method="post">
  <input type="password" name="clave" placeholder="Contraseña" required>
  <button type="submit">Entrar</button>
</form></div></body></html>"""

PANEL = """<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<title>Carnetización - Panel</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#f0f2f5;padding:20px 16px}
.wrap{max-width:700px;margin:0 auto}
.top{display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;flex-wrap:wrap;gap:8px}
.card{background:white;padding:24px;border-radius:16px;box-shadow:0 4px 16px rgba(0,0,0,.1);margin-bottom:18px;text-align:center}
h1{color:#1a1a2e;font-size:19px;margin-bottom:6px}
img.logo{width:60px;margin-bottom:8px}
img.qr{width:170px;height:170px;margin:8px 0}
.url{font-size:11px;color:#aaa;word-break:break-all}
.stats{display:flex;gap:10px;margin-bottom:18px;flex-wrap:wrap}
.stat{flex:1;min-width:80px;background:white;border-radius:12px;padding:12px;text-align:center;box-shadow:0 4px 16px rgba(0,0,0,.08)}
.stat-num{font-size:26px;font-weight:700;color:#1a1a2e}
.stat-label{font-size:11px;color:#888;margin-top:4px}
.tabs{display:flex;gap:6px;margin-bottom:14px;flex-wrap:wrap}
.tab{flex:1;min-width:80px;padding:8px 4px;border-radius:8px;border:2px solid #ddd;cursor:pointer;font-size:12px;background:white;color:#555;font-weight:600}
.tab.active{background:#1a1a2e;color:white;border-color:#1a1a2e}
.section{display:none}.section.active{display:block}
table{width:100%;border-collapse:collapse;font-size:12px;text-align:left}
th{background:#f4f4f4;padding:7px 7px;color:#333}
td{padding:7px 7px;border-bottom:1px solid #eee;vertical-align:middle}
.btn{display:inline-block;padding:4px 9px;border-radius:6px;text-decoration:none;font-size:11px;font-weight:600;border:none;cursor:pointer;white-space:nowrap}
.bl{background:#1a1a2e;color:white}.bg{background:#27ae60;color:white}
.br{background:#e74c3c;color:white}.bo{background:#e67e22;color:white}
.bb{background:#1a73e8;color:white}.bpurple{background:#8e44ad;color:white}
.btn-logout{background:#555;color:white;font-size:11px;padding:6px 12px;border-radius:6px;text-decoration:none}
.badge{display:inline-block;background:#1a1a2e;color:white;border-radius:50%;width:24px;height:24px;line-height:24px;text-align:center;font-weight:700;font-size:11px}
.badge-g{background:#27ae60}
.empty{color:#999;padding:14px 0;font-size:13px;text-align:center}
.nuevo{background:#fff8e1}
.btn-full{width:100%;padding:9px;font-size:13px;margin-top:8px}
.buscar{width:100%;padding:8px 12px;border:1px solid #ddd;border-radius:8px;font-size:13px;margin-bottom:10px;box-sizing:border-box}
.preview-thumb{width:36px;height:36px;object-fit:cover;border-radius:4px;cursor:pointer;border:1px solid #eee}
.modal{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.85);z-index:999;align-items:center;justify-content:center}
.modal.open{display:flex}.modal img{max-width:90%;max-height:90vh;border-radius:8px}
.modal-close{position:absolute;top:16px;right:24px;color:white;font-size:30px;cursor:pointer;font-weight:700}
.bar-wrap{display:flex;align-items:flex-end;gap:8px;height:120px;padding:0 8px}
.bar-col{display:flex;flex-direction:column;align-items:center;flex:1}
.bar{background:#1a73e8;border-radius:4px 4px 0 0;width:100%;min-height:4px;transition:height .3s}
.bar-label{font-size:10px;color:#888;margin-top:4px}
.bar-num{font-size:11px;color:#1a1a2e;font-weight:700;margin-bottom:2px}
.form-row{display:flex;gap:10px;margin-bottom:10px;flex-wrap:wrap}
.form-row input,.form-row select{flex:1;padding:9px;border:1px solid #ddd;border-radius:8px;font-size:13px;min-width:120px}
.form-row button{padding:9px 16px;border:none;border-radius:8px;font-size:13px;cursor:pointer;font-weight:600}
.aviso{background:#fff3cd;border-left:4px solid #f0ad4e;padding:10px 14px;border-radius:6px;font-size:13px;color:#555;margin-bottom:14px;text-align:left}
</style>
<audio id="alerta" src="https://www.soundjay.com/buttons/sounds/button-09.mp3" preload="auto"></audio>
</head><body><div class="wrap">

<div id="modal" class="modal" onclick="cerrarModal()">
  <span class="modal-close">✕</span>
  <img id="modal-img" src="" alt="Vista previa">
</div>

<div class="top">
  <div style="display:flex;align-items:center;gap:10px">
    <img src="{{ logo }}" style="height:36px" onerror="this.style.display='none'">
    <span style="font-size:13px;color:#555">Bienvenida 👋</span>
  </div>
  <div style="display:flex;gap:6px;align-items:center;flex-wrap:wrap">
    <a href="/pantalla" target="_blank" class="btn bpurple">📺 Turnos</a>
    <a href="/exportar" class="btn bg">📥 Excel</a>
    <a href="/logout" class="btn-logout">Salir</a>
  </div>
</div>

<div class="stats">
  <div class="stat"><div class="stat-num">{{ total }}</div><div class="stat-label">Pendientes</div></div>
  <div class="stat"><div class="stat-num" style="color:#27ae60">{{ atendidos_hoy }}</div><div class="stat-label">Atendidos hoy</div></div>
  <div class="stat"><div class="stat-num" style="color:#8e44ad">{% if turno_actual > 0 %}#{{ turno_actual }}{% else %}—{% endif %}</div><div class="stat-label">En pantalla</div></div>
</div>

{% if mensaje_dia %}
<div class="aviso">📢 <b>Aviso activo:</b> {{ mensaje_dia }}</div>
{% endif %}

<div class="card">
  <img class="logo" src="{{ logo }}" onerror="this.style.display='none'">
  <h1>Carnetización UTESA</h1>
  <p style="color:#888;font-size:13px;margin-bottom:8px">Los estudiantes escanean este QR para enviar documentos</p>
  <img class="qr" src="data:image/png;base64,{{ qr }}">
  <p class="url">{{ url_subida }}</p>
</div>

<div class="card" style="text-align:left">
  <div class="tabs">
    <button class="tab active" onclick="mostrar('pendientes',this)">📂 Pendientes ({{ total }})</button>
    <button class="tab" onclick="mostrar('historial',this)">✅ Historial ({{ hist_total }})</button>
    <button class="tab" onclick="mostrar('estadisticas',this)">📊 Estadísticas</button>
    <button class="tab" onclick="mostrar('config',this)">⚙️ Config</button>
  </div>

  <div id="pendientes" class="section active">
    {% if archivos %}
    <table>
      <tr><th>#</th><th>👁</th><th>Nombre</th><th>Hora</th><th>Tam.</th><th colspan="3">Acciones</th></tr>
      {% for a in archivos %}
      <tr class="{{ 'nuevo' if loop.first else '' }}">
        <td><span class="badge">{{ a.turno }}</span></td>
        <td>{% if a.es_imagen %}<img class="preview-thumb" src="/preview/{{ a.archivo }}" onclick="abrirModal('/preview/{{ a.archivo }}')">{% else %}📄{% endif %}</td>
        <td style="word-break:break-all;max-width:110px">{{ a.nombre_estudiante }}</td>
        <td style="white-space:nowrap">{{ a.hora }}</td>
        <td>{{ a.tamano }}</td>
        <td><form method="post" action="/llamar/{{ a.turno }}" style="display:inline"><button class="btn bpurple">📺</button></form></td>
        <td><form method="post" action="/atender/{{ a.archivo }}" style="display:inline" onsubmit="return confirm('¿Marcar atendido?')"><button class="btn bg">✓</button></form></td>
        <td><a class="btn bb" href="/descargar/{{ a.archivo }}">⬇</a></td>
      </tr>
      {% endfor %}
    </table>
    <form method="post" action="/borrar-todos" onsubmit="return confirm('¿Borrar TODOS los pendientes?')" style="margin-top:10px">
      <button class="btn br btn-full">🗑 Borrar todos</button>
    </form>
    <form method="post" action="/reiniciar-turno" onsubmit="return confirm('¿Reiniciar turnos desde #1?')" style="margin-top:8px">
      <button class="btn bo btn-full">🔄 Reiniciar turnos</button>
    </form>
    {% else %}
    <p class="empty">No hay documentos pendientes.</p>
    <form method="post" action="/reiniciar-turno" onsubmit="return confirm('¿Reiniciar turnos desde #1?')" style="margin-top:8px">
      <button class="btn bo btn-full">🔄 Reiniciar turnos</button>
    </form>
    {% endif %}
  </div>

  <div id="historial" class="section">
    <input class="buscar" type="text" id="buscar" placeholder="🔍 Buscar por nombre, fecha o turno..." oninput="filtrarHistorial()">
    {% if historial %}
    <table id="tabla-historial">
      <tr><th>#</th><th>Nombre</th><th>Fecha y hora</th><th>Tamaño</th><th>Estado</th></tr>
      {% for h in historial|reverse %}
      <tr class="fila-hist">
        <td><span class="badge badge-g">{{ h.turno }}</span></td>
        <td style="word-break:break-all;max-width:160px">{{ h.nombre }}</td>
        <td style="white-space:nowrap">{{ h.fecha }}</td>
        <td>{{ h.tamano }}</td>
        <td style="color:#27ae60;font-weight:600">✓ Atendido</td>
      </tr>
      {% endfor %}
    </table>
    {% else %}<p class="empty">El historial está vacío.</p>{% endif %}
  </div>

  <div id="estadisticas" class="section">
    <h3 style="color:#1a1a2e;font-size:15px;margin-bottom:16px">Últimos 7 días</h3>
    <div class="bar-wrap">
      {% set max_val = stats|map(attribute='count')|max if stats else 1 %}
      {% for d in stats %}
      {% set h = (d.count / (max_val if max_val > 0 else 1) * 100)|int %}
      <div class="bar-col">
        <div class="bar-num">{{ d.count }}</div>
        <div class="bar" style="height:{{ h if h > 0 else 4 }}px"></div>
        <div class="bar-label">{{ d.fecha }}</div>
      </div>
      {% endfor %}
    </div>
    <p style="color:#888;font-size:12px;margin-top:16px;text-align:center">Total semana: <b>{{ stats|sum(attribute='count') }}</b> atendidos</p>
  </div>

  <div id="config" class="section">
    <h3 style="color:#1a1a2e;font-size:15px;margin-bottom:16px">⚙️ Configuración</h3>
    <form method="post" action="/cambiar-clave" style="margin-bottom:20px">
      <p style="color:#555;font-size:13px;margin-bottom:10px;text-align:left"><b>Cambiar contraseña</b></p>
      <div class="form-row">
        <input type="password" name="clave_actual" placeholder="Contraseña actual" required>
        <input type="password" name="clave_nueva" placeholder="Nueva contraseña" required>
        <button type="submit" style="background:#1a1a2e;color:white">Cambiar</button>
      </div>
      {% if msg_clave %}<p style="color:{{ 'green' if ok_clave else 'red' }};font-size:13px">{{ msg_clave }}</p>{% endif %}
    </form>
    <form method="post" action="/mensaje-dia">
      <p style="color:#555;font-size:13px;margin-bottom:10px;text-align:left"><b>Mensaje del día</b> (aparece en el formulario de los estudiantes)</p>
      <div class="form-row">
        <input type="text" name="mensaje" placeholder="Ej: Hoy atendemos hasta las 3:00 PM" value="{{ mensaje_dia }}">
        <button type="submit" style="background:#1a73e8;color:white">Guardar</button>
      </div>
      <div class="form-row">
        <button type="submit" name="mensaje" value="" style="background:#e74c3c;color:white;width:100%">🗑 Borrar mensaje</button>
      </div>
    </form>
  </div>

</div>
</div>
<script>
function mostrar(id,btn){
  document.querySelectorAll('.section').forEach(s=>s.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  btn.classList.add('active');
}
function abrirModal(src){document.getElementById('modal-img').src=src;document.getElementById('modal').classList.add('open');}
function cerrarModal(){document.getElementById('modal').classList.remove('open');}
function filtrarHistorial(){
  var q=document.getElementById('buscar').value.toLowerCase();
  document.querySelectorAll('.fila-hist').forEach(function(r){r.style.display=r.textContent.toLowerCase().includes(q)?'':'none';});
}
// Cerrar sesión al cerrar la pestaña
window.addEventListener('beforeunload', function(){
  navigator.sendBeacon('/logout-beacon');
});
var totalAnterior={{ total }};
setInterval(function(){
  fetch('/total').then(r=>r.json()).then(d=>{
    if(d.total>totalAnterior){document.getElementById('alerta').play();totalAnterior=d.total;location.reload();}
  });
},5000);
</script></body></html>"""

FORM = """<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Enviar documento</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#f0f2f5;padding:24px 16px}
.card{background:white;max-width:420px;margin:0 auto;padding:24px;border-radius:16px;box-shadow:0 4px 16px rgba(0,0,0,.1)}
.header{text-align:center;margin-bottom:18px}
.header img{width:60px;margin-bottom:8px}
h2{color:#1a1a2e;font-size:19px}
label{display:block;margin:14px 0 5px;font-size:14px;color:#444;font-weight:600}
input,select{width:100%;padding:10px;border:1px solid #ccc;border-radius:8px;font-size:15px}
button{width:100%;margin-top:18px;padding:14px;background:#1a73e8;color:white;border:none;border-radius:8px;font-size:16px;cursor:pointer}
.info{font-size:12px;color:#888;margin-top:4px}
.aviso{background:#fff3cd;border-left:4px solid #f0ad4e;padding:10px 14px;border-radius:6px;font-size:13px;color:#555;margin-bottom:14px}
.err-size{color:#c0392b;font-size:13px;margin-top:6px;display:none}
.preview-box{margin-top:10px;display:none}
.preview-box img{max-width:100%;max-height:180px;border-radius:8px;border:1px solid #eee}
.file-list{margin-top:8px;font-size:12px;color:#555}
</style></head><body><div class="card">
<div class="header">
  <img src="{{ logo }}" onerror="this.style.display='none'">
  <h2>Enviar documento</h2>
</div>
{% if mensaje_dia %}<div class="aviso">📢 {{ mensaje_dia }}</div>{% endif %}
<form method="post" enctype="multipart/form-data" onsubmit="return validar()">
  <label>Tu nombre</label>
  <input type="text" name="nombre" placeholder="Ej. Juan Pérez" required>
  <label>Selecciona los documentos</label>
  <input type="file" name="docs" id="docs" accept=".pdf,.jpg,.jpeg,.png,.doc,.docx" multiple required onchange="mostrarPreview()">
  <div class="info">Formatos: PDF, Word, JPG, PNG &nbsp;•&nbsp; Máx. 25 MB por archivo &nbsp;•&nbsp; Puedes seleccionar varios</div>
  <div id="err-size" class="err-size">⚠ Uno o más archivos superan 25 MB.</div>
  <div class="preview-box" id="preview-box">
    <img id="preview-img" src="" style="display:none">
    <div class="file-list" id="file-list"></div>
  </div>
  <button type="submit">📤 Enviar documento(s)</button>
</form></div>
<script>
function mostrarPreview(){
  var files=document.getElementById('docs').files;
  var err=document.getElementById('err-size');
  var box=document.getElementById('preview-box');
  var list=document.getElementById('file-list');
  var img=document.getElementById('preview-img');
  var hayError=false;
  var html='<b>Archivos seleccionados:</b><ul style="margin:6px 0 0 16px">';
  Array.from(files).forEach(function(f){
    var mb=(f.size/1024/1024).toFixed(2);
    if(f.size>25*1024*1024) hayError=true;
    html+='<li>'+f.name+' ('+mb+' MB)</li>';
    if(f.type.startsWith('image/') && files.length===1){
      var reader=new FileReader();
      reader.onload=function(e){img.src=e.target.result;img.style.display='block';};
      reader.readAsDataURL(f);
    }
  });
  html+='</ul>';
  list.innerHTML=html;
  box.style.display='block';
  err.style.display=hayError?'block':'none';
}
function validar(){
  var files=document.getElementById('docs').files;
  for(var i=0;i<files.length;i++){if(files[i].size>25*1024*1024){document.getElementById('err-size').style.display='block';return false;}}
  return true;
}
</script></body></html>"""

CONFIRMACION = """<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Enviado</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#f0f2f5;padding:40px 16px;text-align:center}
.card{background:white;max-width:400px;margin:0 auto;padding:32px 24px;border-radius:16px;box-shadow:0 4px 16px rgba(0,0,0,.1)}
img.logo{width:60px;margin-bottom:10px}
.turno{font-size:90px;font-weight:700;color:#1a1a2e;line-height:1;margin:12px 0}
h2{color:#1a1a2e;font-size:19px;margin-bottom:6px}
.ok{color:#27ae60;font-size:24px;margin-bottom:6px}
.detalle{background:#f4f6f9;border-radius:8px;padding:12px;margin-top:16px;font-size:13px;color:#555;text-align:left;line-height:2}
.detalle b{color:#1a1a2e}
</style></head><body><div class="card">
<img class="logo" src="{{ logo }}" onerror="this.style.display='none'">
<div class="ok">✅</div>
<h2>¡Documento(s) enviado(s)!</h2>
<p style="color:#666;font-size:14px">Tu número de turno es:</p>
<div class="turno">#{{ turno }}</div>
<p style="color:#888;font-size:13px">Espera a que te llamen</p>
<div class="detalle">
  <div><b>Nombre:</b> {{ nombre }}</div>
  <div><b>Archivos:</b> {{ archivos }}</div>
  <div><b>Hora:</b> {{ fecha }}</div>
</div>
</div></body></html>"""

PANTALLA = """<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<title>Turnos - Carnetización</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#1a1a2e;display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:100vh;text-align:center;padding:20px}
img.logo{width:80px;margin-bottom:12px;opacity:0.9}
h1{color:white;font-size:24px;margin-bottom:4px;opacity:0.8}
.sub{color:#aaa;font-size:14px;margin-bottom:32px}
.turno-box{background:white;border-radius:24px;padding:32px 52px;margin-bottom:24px}
.label{color:#888;font-size:16px;margin-bottom:8px}
.numero{font-size:140px;font-weight:700;color:#1a1a2e;line-height:1}
.esperando{color:#aaa;font-size:20px;margin-top:10px}
.atendidos{background:rgba(255,255,255,0.1);border-radius:12px;padding:12px 28px;color:white}
.at-num{font-size:40px;font-weight:700;color:#27ae60}
.at-label{font-size:12px;color:#aaa;margin-top:4px}
</style></head><body>
<img class="logo" src="{{ logo }}" onerror="this.style.display='none'">
<h1>Carnetización UTESA</h1>
<p class="sub">Pantalla de turnos</p>
<div class="turno-box">
  <div class="label">Turno en atención</div>
  {% if turno > 0 %}<div class="numero">#{{ turno }}</div>
  {% else %}<div class="esperando">Esperando...</div>{% endif %}
</div>
<div class="atendidos">
  <div class="at-num">{{ atendidos }}</div>
  <div class="at-label">Atendidos hoy</div>
</div>
<script>setInterval(function(){location.reload();},4000);</script>
</body></html>"""

# ─── RUTAS ───────────────────────────────────────────────────

@app.route("/", methods=["GET","POST"])
def login():
    if session.get("auth"):
        return redirect("/panel")
    if request.method == "POST":
        if request.form.get("clave") == get_clave():
            session["auth"] = True
            return redirect("/panel")
        return render_template_string(LOGIN, error=True, logo=UTESA_LOGO)
    return render_template_string(LOGIN, error=False, logo=UTESA_LOGO)

@app.route("/panel")
def panel():
    if not session.get("auth"):
        return redirect("/")
    host = request.host_url.rstrip("/")
    url_subida = host + "/subir"
    archivos_raw = sorted(os.listdir(CARPETA), reverse=True)
    archivos = []
    for nombre in archivos_raw:
        ruta = os.path.join(CARPETA, nombre)
        stat = os.stat(ruta)
        mb = stat.st_size/1024/1024
        tamano = f"{mb:.2f} MB" if mb>=1 else f"{stat.st_size//1024} KB"
        partes = nombre.split("_")
        hora = turno_num = nombre_est = ""
        try:
            hora = datetime.strptime(partes[0]+"_"+partes[1], "%Y%m%d_%H%M%S").strftime("%H:%M")
            turno_num = partes[2]
            nombre_est = partes[3].replace("_"," ") if len(partes)>3 else nombre
        except:
            nombre_est = nombre
        archivos.append({"archivo":nombre,"nombre_estudiante":nombre_est,"hora":hora,
                         "tamano":tamano,"turno":turno_num,"es_imagen":es_imagen(nombre)})
    historial = cargar_historial()
    e = get_estado()
    cfg = cargar_config()
    return render_template_string(PANEL, qr=generar_qr(url_subida), url_subida=url_subida,
                                  archivos=archivos, total=len(archivos),
                                  historial=historial, hist_total=len(historial),
                                  atendidos_hoy=e["atendidos_hoy"],
                                  turno_actual=e.get("turno_actual",0),
                                  stats=stats_semana(),
                                  mensaje_dia=cfg.get("mensaje_dia",""),
                                  msg_clave=session.pop("msg_clave",None),
                                  ok_clave=session.pop("ok_clave",False),
                                  logo=UTESA_LOGO)

@app.route("/total")
def total():
    return {"total": len(os.listdir(CARPETA))}

@app.route("/preview/<nombre>")
def preview(nombre):
    if not session.get("auth"):
        return redirect("/")
    ruta = os.path.join(CARPETA, nombre)
    if os.path.exists(ruta) and es_imagen(nombre):
        return send_file(ruta)
    return "No disponible", 404

@app.route("/subir", methods=["GET","POST"])
def subir():
    cfg = cargar_config()
    if request.method == "POST":
        files = request.files.getlist("docs")
        nombre = request.form.get("nombre","").strip() or "Estudiante"
        if not files or all(f.filename=="" for f in files):
            return render_template_string(FORM, logo=UTESA_LOGO, mensaje_dia=cfg.get("mensaje_dia",""))
        turno = siguiente_turno()
        fecha = datetime.now()
        fecha_str = fecha.strftime("%d/%m/%Y %H:%M")
        marca = fecha.strftime("%Y%m%d_%H%M%S")
        nombre_limpio = nombre.replace(" ","_")
        nombres_archivos = []
        for i, f in enumerate(files):
            if f.filename == "" or not ext_ok(f.filename):
                continue
            contenido = f.read()
            if len(contenido) > MAX_MB*1024*1024:
                continue
            sufijo = f"_{i}" if i > 0 else ""
            nombre_final = f"{marca}_{turno}_{nombre_limpio}{sufijo}_{f.filename.replace(' ','_')}"
            with open(os.path.join(CARPETA, nombre_final),"wb") as out:
                out.write(contenido)
            nombres_archivos.append(f.filename)
        if not nombres_archivos:
            return render_template_string(FORM, logo=UTESA_LOGO, mensaje_dia=cfg.get("mensaje_dia",""))
        h = cargar_historial()
        h.append({"turno":turno,"nombre":nombre,"archivo":", ".join(nombres_archivos),"fecha":fecha_str,"tamano":f"{len(nombres_archivos)} archivo(s)"})
        guardar_historial(h)
        return render_template_string(CONFIRMACION, turno=turno, nombre=nombre,
                                      archivos=", ".join(nombres_archivos), fecha=fecha_str, logo=UTESA_LOGO)
    return render_template_string(FORM, logo=UTESA_LOGO, mensaje_dia=cfg.get("mensaje_dia",""))

@app.route("/exportar")
def exportar():
    if not session.get("auth"):
        return redirect("/")
    historial = cargar_historial()
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    # Colores
    azul_oscuro = "1A1A2E"
    azul_claro  = "D6E4F7"
    verde       = "27AE60"
    blanco      = "FFFFFF"
    gris        = "F4F4F4"

    # Fila 1: Título principal
    ws.merge_cells("A1:F1")
    titulo = ws["A1"]
    titulo.value = "UNIVERSIDAD UTESA — Departamento de Carnetización"
    titulo.font = Font(bold=True, size=14, color=blanco)
    titulo.fill = PatternFill("solid", fgColor=azul_oscuro)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Fila 2: Fecha de exportación
    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"Exportado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub.font = Font(italic=True, size=10, color="555555")
    sub.fill = PatternFill("solid", fgColor="EEF3FF")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Fila 3: vacía
    ws.row_dimensions[3].height = 6

    # Fila 4: Encabezados de columnas
    headers = ["# Turno", "Nombre del estudiante", "Fecha y hora", "Archivos enviados", "Tamaño", "Estado"]
    borde = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, size=11, color=blanco)
        cell.fill = PatternFill("solid", fgColor=azul_oscuro)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    ws.row_dimensions[4].height = 22

    # Filas de datos
    for i, h in enumerate(historial):
        row = i + 5
        fila_color = blanco if i % 2 == 0 else gris
        valores = [
            h.get("turno", ""),
            h.get("nombre", ""),
            h.get("fecha", ""),
            h.get("archivo", ""),
            h.get("tamano", ""),
            "✓ Atendido"
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=str(val))
            cell.fill = PatternFill("solid", fgColor=fila_color)
            cell.border = borde
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)
            if col == 6:
                cell.font = Font(color=verde, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 18

    # Fila resumen
    total_row = len(historial) + 5
    ws.merge_cells(f"A{total_row}:E{total_row}")
    resumen = ws[f"A{total_row}"]
    resumen.value = f"Total de estudiantes atendidos: {len(historial)}"
    resumen.font = Font(bold=True, size=11, color=azul_oscuro)
    resumen.fill = PatternFill("solid", fgColor=azul_claro)
    resumen.alignment = Alignment(horizontal="right", vertical="center")
    ws[f"F{total_row}"].fill = PatternFill("solid", fgColor=azul_claro)
    ws.row_dimensions[total_row].height = 20

    # Anchos de columna
    anchos = [10, 28, 18, 35, 12, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fecha_hoy = datetime.now().strftime("%Y%m%d")
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename=historial_carnetizacion_{fecha_hoy}.xlsx"}
    )

@app.route("/cambiar-clave", methods=["POST"])
def cambiar_clave():
    if not session.get("auth"):
        return redirect("/")
    cfg = cargar_config()
    actual = request.form.get("clave_actual","")
    nueva = request.form.get("clave_nueva","")
    if actual == cfg.get("clave", CLAVE_DEFAULT):
        cfg["clave"] = nueva
        guardar_config(cfg)
        session["msg_clave"] = "✅ Contraseña cambiada correctamente."
        session["ok_clave"] = True
    else:
        session["msg_clave"] = "❌ La contraseña actual es incorrecta."
        session["ok_clave"] = False
    return redirect("/panel")

@app.route("/mensaje-dia", methods=["POST"])
def mensaje_dia():
    if not session.get("auth"):
        return redirect("/")
    cfg = cargar_config()
    cfg["mensaje_dia"] = request.form.get("mensaje","").strip()
    guardar_config(cfg)
    return redirect("/panel")

@app.route("/llamar/<int:turno>", methods=["POST"])
def llamar(turno):
    if not session.get("auth"):
        return redirect("/")
    e = get_estado()
    e["turno_actual"] = turno
    guardar_estado(e)
    return redirect("/panel")

@app.route("/atender/<nombre>", methods=["POST"])
def atender(nombre):
    if not session.get("auth"):
        return redirect("/")
    ruta = os.path.join(CARPETA, nombre)
    if os.path.exists(ruta):
        partes = nombre.split("_")
        turno_num = int(partes[2]) if len(partes)>2 and partes[2].isdigit() else 0
        nombre_est = partes[3].replace("_"," ") if len(partes)>3 else nombre
        hora = ""
        try:
            hora = datetime.strptime(partes[0]+"_"+partes[1], "%Y%m%d_%H%M%S").strftime("%d/%m/%Y %H:%M")
        except:
            hora = datetime.now().strftime("%d/%m/%Y %H:%M")
        stat = os.stat(ruta)
        mb = stat.st_size/1024/1024
        tamano = f"{mb:.2f} MB" if mb>=1 else f"{stat.st_size//1024} KB"
        h = cargar_historial()
        h.append({"turno":turno_num,"nombre":nombre_est,"archivo":nombre,"fecha":hora,"tamano":tamano})
        guardar_historial(h)
        e = get_estado()
        e["atendidos_hoy"] = e.get("atendidos_hoy",0) + 1
        guardar_estado(e)
        os.remove(ruta)
    return redirect("/panel")

@app.route("/descargar/<nombre>")
def descargar(nombre):
    if not session.get("auth"):
        return redirect("/")
    ruta = os.path.join(CARPETA, nombre)
    if os.path.exists(ruta):
        return send_file(ruta, as_attachment=True)
    return "Archivo no encontrado.", 404

@app.route("/borrar/<nombre>", methods=["POST"])
def borrar(nombre):
    if not session.get("auth"):
        return redirect("/")
    ruta = os.path.join(CARPETA, nombre)
    if os.path.exists(ruta):
        os.remove(ruta)
    return redirect("/panel")

@app.route("/borrar-todos", methods=["POST"])
def borrar_todos():
    if not session.get("auth"):
        return redirect("/")
    for f in os.listdir(CARPETA):
        os.remove(os.path.join(CARPETA, f))
    return redirect("/panel")

@app.route("/reiniciar-turno", methods=["POST"])
def reiniciar_turno():
    if not session.get("auth"):
        return redirect("/")
    with open(TURNO_FILE,"w") as f:
        json.dump({"actual":0}, f)
    return redirect("/panel")

@app.route("/pantalla")
def pantalla():
    e = get_estado()
    return render_template_string(PANTALLA, turno=e.get("turno_actual",0),
                                  atendidos=e.get("atendidos_hoy",0), logo=UTESA_LOGO)

@app.route("/logout-beacon", methods=["POST"])
def logout_beacon():
    session.clear()
    return "", 204

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

if __name__ == "__main__":
    port = int(os.environ.get("PORT",5000))
    app.run(host="0.0.0.0", port=port)
